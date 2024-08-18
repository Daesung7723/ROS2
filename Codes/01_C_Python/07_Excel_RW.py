import pandas as pd
from openpyxl import load_workbook
import os
os.chdir(os.path.dirname(os.path.abspath(__file__)))

d1 = ["kim", "dae", "sung"]
d2 = ["1977", "2000", "2024"]
d3 = ["1", "23", "47"]
d4 = ['22','33','55']
data = [d1,d2,d3,d4]
df1 = pd.DataFrame(data)

def Excel_Save(save_data, filename) :
    print(save_data)
    save_data.to_excel(filename, index=False, header=False)


def Excel_Load_reform(filename) :
    load_wb = load_workbook(filename)
    load_ws = load_wb.active

    # read_data = [[0 for i in range(load_ws.max_row)] for j in range(load_ws.max_column)]
    read_data = [[0 for i in range(load_ws.max_row)] for j in range(load_ws.max_column)]

    for r in range(1, load_ws.max_column+1):
        for c in range(1, load_ws.max_row+1):
            read_data[r-1][c-1] = load_ws.cell(c,r).value

    df = pd.DataFrame(read_data)
    return df

if __name__ == '__main__' :
    Excel_Save(df1, 'file1.xlsx')
    rtn_df = Excel_Load_reform('file1.xlsx')
    Excel_Save(rtn_df, 'file2.xlsx')